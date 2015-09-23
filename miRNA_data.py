import urllib2
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from easygui import *
msgbox("Select the directory in which excel sheet will be saved or is already present")
d=diropenbox()
d=d+'\\data.xlsx'
initial=['ID','Accession','Stem-loop Sequence','Location','FASTA Sequence']
#print d
try:
    wb=load_workbook(filename=d)
    ws=wb.worksheets[0]
    row_count=ws.get_highest_row()-1
except:
    wb=Workbook()
    ws=wb.active
    ws.append(initial)
    row_count=0
    d='data.xlsx'
#base_url='test.html'
#html=urllib2.urlopen(base_url).read()
html=open('test.html','r')
soup=BeautifulSoup(html)
f=open('html.txt','w')
table=soup.find('table',attrs={'class':'resultTable'}).tbody
rows=table.find_all('tr')
total_rows=len(rows)
base='http://www.ncbi.nlm.nih.gov/'
form='http://www.ncbi.nlm.nih.gov/sviewer/viewer.fcgi?val=%s&db=nuccore&dopt=fasta&extrafeat=0&fmt_mask=0&%s&retmode=html&withmarkup=on&log$=seqview&maxdownloadsize=1000000'
print 'Starting to Download...'
for i in range(row_count,total_rows):
    x=rows[i].td
    link=x.a['href']
    title=str(x.a.string)
    t=link.index('acc=')
    access=link[t+4:]
    s='http://www.mirbase.org/cgi-bin/get_seq.pl?acc='+access
    h=urllib2.urlopen(s).read()[:-8]
    m=h.rfind('\n')
    stem=h[m+1:]
    title=str(x.a.string)
    g=urllib2.urlopen(link).read()
    x=g.index('ENTREZGENE: ')
    y=g.index(';',x)
    t=g[x+12:y]
    while True:
        try:
            soup=BeautifulSoup(urllib2.urlopen(base+'gene/'+t).read())
            break
        except:
            continue
    l=str(soup.find('div',attrs={'class':'gt_cont_contents'}).dd.span.string)
    fasta=(soup.find('div',attrs={'class':'seq-viewer'}).find_all('a')[1])['href']
    o=fasta.index('from')
    to_from=fasta[o:]
    total=BeautifulSoup(urllib2.urlopen(base+fasta).read()).find('div',attrs={'id':'viewercontent1','class':'seq gbff'})['val']
    final=form %(total, to_from)
    soup=BeautifulSoup(urllib2.urlopen(final).read())
    span=soup.find_all('span')
    total=len(span)
    final=''
    for j in range(total):
        span[j]=str(span[j].string)
        final+=span[j]
    info=[title,access,stem,l,final]
    ws.append(info)
    wb.save(d)
    print str(i+1)+'th done'
